from datetime import datetime

import openpyxl

EXCEL_FILE = "passwords.xlsx"


def initialize_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Passwords"
        sheet.append(["ID", "Password", "Security Level", "Created At"])
        wb.save(EXCEL_FILE)


def read_passwords():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    return [
        {
            "id": row[0].value,
            "password": row[1].value,
            "security_level": row[2].value,
            "created_at": row[3].value,
        }
        for row in sheet.iter_rows(min_row=2)
    ]


def add_password(password, security_level):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    new_id = sheet.max_row
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([new_id, password, security_level, created_at])
    wb.save(EXCEL_FILE)


def delete_password(password_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == password_id:
            sheet.delete_rows(row[0].row)
            wb.save(EXCEL_FILE)
            return True
    return False


initialize_excel()
